"""Core logic for aggregating telecom invoices."""

from __future__ import annotations

import io
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, MutableMapping, Tuple, Union

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PathLike = Union[str, Path]


@dataclass
class InvoiceBreakdown:
    """Intermediate representation for a single invoice."""

    mois: str
    date: pd.Timestamp
    numero_facture: int
    montants: Dict[str, float]
    total: float
    total_ttc: float
    nb_lignes_detail: int
    statut: str = "Validee"

    def to_dict(self) -> Dict[str, object]:
        """Return a serializable representation."""

        return {
            "mois": self.mois,
            "date": self.date,
            "numero_facture": self.numero_facture,
            "montants": dict(self.montants),
            "total": self.total,
            "total_ttc": self.total_ttc,
            "nb_lignes_detail": self.nb_lignes_detail,
            "statut": self.statut,
        }


class TelecomInvoiceProcessor:
    """Load and aggregate telecom invoice CSV exports."""

    MONTANT_COLS = [
        "Montant (€ HT)",
        "Montant (€ TTC)",
        "Montant (€'€ HT)",
        "Montant (€'€ TTC)",
        "Montant (? HT)",
        "Montant (? TTC)",
        "Montant (�'� HT)",
        "Montant (�'� TTC)",
    ]
    COLUMN_ALIASES: Dict[str, List[str]] = {
        "numero_compte": ["Numéro compte", "Num�ro compte"],
        "numero_facture": ["Numéro facture", "Num�ro facture"],
        "date": ["Date"],
        "niveau_charge": ["Niveau de charge"],
        "type_acces": ["Type d'accès", "Type d'acc�s"],
        "numero_acces": ["Numéro accès", "Num�ro acc�s", "Numéro acc�s"],
        "libelle": ["Libellé ligne facture", "Libell� ligne facture"],
        "rubrique_facture": ["Rubrique facture"],
        "indicateur_abo": ["Indicateur abo principal/option"],
        "type_charge": ["Type de charge"],
        "type_trafic": ["Type de trafic"],
        "unite": ["Unité", "Unit�"],
        "quantite_acte": ["Quantité (Acte)", "Quantit� (Acte)"],
        "nombre_lignes": ["Nombre de lignes"],
        "quantite_volume_mo": ["Quantité (Volume (Mo))", "Quantit� (Volume (Mo))"],
        "quantite_duree": ["Quantité (Durée (hh:mm:ss))", "Quantit� (Dur�e (hh:mm:ss))"],
        "quantite_volume_ko": ["Quantité (Volume (Ko))", "Quantit� (Volume (Ko))"],
        "quantite_hors_conso": ["Quantité (Hors conso)", "Quantit� (Hors conso)"],
        "montant_ht": [
            "Montant (€ HT)",
            "Montant (€'€ HT)",
            "Montant (? HT)",
            "Montant (� HT)",
            "Montant (�'� HT)",
        ],
        "montant_tva": ["Montant TVA"],
        "montant_ttc": [
            "Montant (€ TTC)",
            "Montant (€'€ TTC)",
            "Montant (? TTC)",
            "Montant (� TTC)",
            "Montant (�'� TTC)",
        ],
    }
    FILTER_FIELDS = {
        "date",
        "date_debut",
        "date_fin",
        "numero_compte",
        "numero_facture",
        "numero_acces",
        "type_acces",
        "type_ligne",
        "type_charge",
    }

    def __init__(self) -> None:
        self.data: pd.DataFrame | None = None
        self.aggregated_data: Dict[str, Dict[str, object]] | None = None
        self.columns: Dict[str, str] = {}
        self.accounts_with_fiber: set[str] = set()

    # ------------------------------------------------------------------ #
    # Data loading and inspection
    def load_csv(self, filepath: PathLike, *, silent: bool = False) -> pd.DataFrame:
        """Load a CSV file exported from Orange Business."""

        self.data = pd.read_csv(filepath, sep=";", encoding="utf-8-sig")
        self._resolve_columns()
        self._normalize_numeric_columns()
        self._mark_accounts_with_fiber()
        if not silent:
            self._print_load_summary()
        return self.data

    def load_csv_content(self, content: bytes, *, silent: bool = True) -> pd.DataFrame:
        """Load CSV content from bytes (useful for uploads or API usage)."""

        self.data = pd.read_csv(io.BytesIO(content), sep=";", encoding="utf-8-sig")
        self._resolve_columns()
        self._normalize_numeric_columns()
        self._mark_accounts_with_fiber()
        if not silent:
            self._print_load_summary()
        return self.data

    def display_structure(self) -> None:
        """Print quick stats about the loaded dataset."""

        if self.data is None:
            print("Aucune donnée chargée.")
            return

        print("\n=== Structure des données ===")
        print(f"Nombre de lignes: {len(self.data)}")
        print(f"Nombre de colonnes: {len(self.data.columns)}")

        rubrique_counts = self.data[self.col("rubrique_facture")].value_counts()
        print("\nRépartition par rubrique:")
        print(rubrique_counts)

        type_counts = self.data[self.col("type_charge")].value_counts()
        print("\nRépartition par type de charge:")
        print(type_counts)

        total_ht = self.data[self.col("montant_ht")].sum()
        total_ttc = self.data[self.col("montant_ttc")].sum()
        print(f"\nMontant total HT: {total_ht:,.2f} €")
        print(f"Montant total TTC: {total_ttc:,.2f} €")

    # ------------------------------------------------------------------ #
    # Aggregation
    def categorize_charge(self, row: pd.Series) -> str:
        """Return the high-level category for a row."""

        type_charge = str(row.get(self.col("type_charge"), "")).lower()
        rubrique = str(row.get(self.col("rubrique_facture"), "")).lower()

        if "remise" in type_charge or "remise" in rubrique:
            return "Remise"
        if "abonnement" in type_charge or "abonnement" in rubrique:
            return "Abo"
        return "Conso"

    def classify_access_type(self, value: str, account: str | int | None = None, libelle: str | None = None) -> str:
        """Categorise access types into broad buckets."""

        label = self._normalize_text(str(value or ""))
        lib_norm = self._normalize_text(str(libelle or ""))
        account_str = str(account) if account is not None else ""

        bas_keywords = ["bas d", "bas debit", "bas-debit", "basdebit", "rtc", "numeris", "t2", "t0", "adsl", "sdsl", "vdsl", "cuivre", "dsl"]
        fiber_keywords = ["fibre", "fiber", "ftth", "ftto", "fttx", "optique"]

        line_is_bas = any(k in label or k in lib_norm for k in bas_keywords)
        line_is_fiber = any(k in label or k in lib_norm for k in fiber_keywords)

        if "mobile" in label or "sim" in label:
            return "Mobile"
        if line_is_fiber:
            return "Internet"
        if "internet" in label or "adsl" in label or "vdsl" in label or line_is_bas:
            return "Internet bas debit"
        if "secours" in label or "secondaire" in label:
            return "Fixe secondaire"
        if "rtc" in label or "numeris" in label or "t2" in label or "t0" in label or "ligne" in label:
            return "Fixe"
        return "Autre"

    def aggregate_by_account(self) -> Dict[str, Dict[int, InvoiceBreakdown]]:
        """Group the dataset by billing account and invoice number."""

        if self.data is None:
            raise ValueError("Aucune donnée chargée. Utilisez load_csv() d'abord.")

        df = self.data.copy()
        df["Categorie"] = df.apply(self.categorize_charge, axis=1)

        aggregates: Dict[str, Dict[int, InvoiceBreakdown]] = {}

        for account, account_group in df.groupby(self.col("numero_compte")):
            invoices: Dict[int, InvoiceBreakdown] = {}
            for facture_num, facture_data in account_group.groupby(self.col("numero_facture")):
                date_obj = pd.to_datetime(facture_data[self.col("date")].iloc[0], format="%d/%m/%Y")
                mois = self._month_label(date_obj)
                montants = {
                    "Abo": facture_data.loc[facture_data["Categorie"] == "Abo", self.col("montant_ht")].sum(),
                    "Conso": facture_data.loc[facture_data["Categorie"] == "Conso", self.col("montant_ht")].sum(),
                    "Remise": facture_data.loc[facture_data["Categorie"] == "Remise", self.col("montant_ht")].sum(),
                }

                invoices[int(facture_num)] = InvoiceBreakdown(
                    mois=mois,
                    date=date_obj,
                    numero_facture=int(facture_num),
                    montants=montants,
                    total=facture_data[self.col("montant_ht")].sum(),
                    total_ttc=facture_data[self.col("montant_ttc")].sum(),
                    nb_lignes_detail=len(facture_data),
                )

            aggregates[str(account)] = {
                "factures": invoices,
                "lignes_telecom": int(account_group[self.col("numero_acces")].nunique()),
            }

        self.aggregated_data = aggregates
        print(f"\n✅ Agrégation terminée: {len(aggregates)} comptes traités")
        return aggregates

    def aggregate_by_account_filtered(
        self,
        *,
        date: str | None = None,
        date_debut: str | None = None,
        date_fin: str | None = None,
        numero_compte: str | None = None,
        numero_facture: int | None = None,
        numero_acces: str | None = None,
        type_acces: List[str] | None = None,
        type_ligne: List[str] | None = None,
        type_charge: List[str] | None = None,
    ) -> Dict[str, Dict[int, InvoiceBreakdown]]:
        """Same aggregation but only on the filtered subset (type_ligne, dates...)."""

        df = self._filtered_dataframe(
            date=date,
            date_debut=date_debut,
            date_fin=date_fin,
            numero_compte=numero_compte,
            numero_facture=numero_facture,
            numero_acces=numero_acces,
            type_acces=type_acces,
            type_ligne=type_ligne,
            type_charge=type_charge,
        )

        # Keep consistent categorisation column name
        df["Categorie"] = df["categorie_charge"]
        self.data = df  # keep internal reference consistent for further exports if needed
        return self.aggregate_by_account()

    # ------------------------------------------------------------------ #
    # Reporting helpers
    def generate_report(self) -> None:
        """Print a textual summary grouped by account."""

        if self.aggregated_data is None:
            print("Aucune donnée agrégée disponible.")
            return

        print("\n" + "=" * 80)
        print("RAPPORT DE FACTURATION TELECOM PAR COMPTE")
        print("=" * 80)

        for account, data in sorted(self.aggregated_data.items()):
            factures = data["factures"]
            print(f"\nCompte: {account}")
            print(f"  Nombre de lignes telecom: {data['lignes_telecom']}")
            print(f"  Nombre de factures: {len(factures)}")
            for facture in factures.values():
                self._print_invoice(facture)

    def export_to_excel_format(self, output_path: PathLike, month: str) -> None:
        """Write an Excel workbook following the sample structure."""

        if self.aggregated_data is None:
            raise ValueError("Aucune donnée agrégée. Utilisez aggregate_by_account().")

        wb = Workbook()
        ws = wb.active
        ws.title = "Lot 1 Fixe"

        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        headers = [
            ("A1", "Compte de facturation"),
            ("B1", "Nom du compte"),
            ("C1", month),
            ("C2", "n° de facture"),
            ("D2", "Résumé"),
            ("E2", "Montant € HT"),
            ("F2", "Statut"),
        ]

        for cell, value in headers:
            ws[cell] = value
            ws[cell].font = header_font
            ws[cell].fill = header_fill

        row = 3
        month_lower = month.lower()

        for account, data in sorted(self.aggregated_data.items()):
            for facture in data["factures"].values():
                if facture.mois.lower() != month_lower:
                    continue
                ws.cell(row, 1, int(account))
                ws.cell(row, 2, "")
                ws.cell(row, 3, facture.numero_facture)
                ws.cell(row, 4, "Total")
                ws.cell(row, 5, round(facture.total, 2))
                ws.cell(row, 6, facture.statut)

                row = self._write_breakdown_rows(ws, row, facture)

        for col in "ABCDEF":
            ws.column_dimensions[col].width = 20
        for cell in ws[1] + ws[2]:
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

        wb.save(output_path)
        print(f"✅ Fichier Excel créé: {output_path}")

    def export_summary_csv(self, output_path: PathLike) -> None:
        """Export the aggregated summary to CSV."""

        if self.aggregated_data is None:
            raise ValueError("Aucune donnée agrégée disponible.")

        rows: List[dict] = []
        for account, data in sorted(self.aggregated_data.items()):
            for facture_num, facture in data["factures"].items():
                rows.append(
                    {
                        "Compte": account,
                        "Facture": facture_num,
                        "Mois": facture.mois,
                        "Abonnements_HT": round(facture.montants["Abo"], 2),
                        "Consommations_HT": round(facture.montants["Conso"], 2),
                        "Remises_HT": round(facture.montants["Remise"], 2),
                        "Total_HT": round(facture.total, 2),
                        "Total_TTC": round(facture.total_ttc, 2),
                        "Nb_lignes_telecom": data["lignes_telecom"],
                        "Statut": facture.statut,
                    }
                )

        pd.DataFrame(rows).to_csv(output_path, index=False, encoding="utf-8-sig", sep=";")
        print(f"✅ Fichier CSV résumé créé: {output_path}")

    def filter_lines(
        self,
        *,
        date: str | None = None,
        date_debut: str | None = None,
        date_fin: str | None = None,
        numero_compte: str | None = None,
        numero_facture: int | None = None,
        numero_acces: str | None = None,
        type_acces: List[str] | None = None,
        type_ligne: List[str] | None = None,
        type_charge: List[str] | None = None,
    ) -> Tuple[List[Dict[str, object]], Dict[str, object]]:
        """Filter raw lines with SQL-like criteria."""

        df = self._filtered_dataframe(
            date=date,
            date_debut=date_debut,
            date_fin=date_fin,
            numero_compte=numero_compte,
            numero_facture=numero_facture,
            numero_acces=numero_acces,
            type_acces=type_acces,
            type_ligne=type_ligne,
            type_charge=type_charge,
        )
        rows: List[Dict[str, object]] = []
        per_type: Dict[str, Dict[str, object]] = {}
        montants = df[self.col("montant_ht")].fillna(0).astype(float)

        for _, row in df.iterrows():
            date_value = row["__date_obj"]
            type_label = row["type_ligne"]
            montant_val = float(row.get(self.col("montant_ht"), 0) or 0)

            per_type.setdefault(type_label, {"count": 0, "montant_ht": 0.0})
            per_type[type_label]["count"] += 1
            per_type[type_label]["montant_ht"] += montant_val

            rows.append(
                {
                    "compte": row[self.col("numero_compte")],
                    "facture": row[self.col("numero_facture")],
                    "date": date_value.isoformat() if pd.notna(date_value) else "",
                    "numero_acces": row[self.col("numero_acces")],
                    "type_acces": row[self.col("type_acces")],
                    "type_ligne": type_label,
                    "categorie_charge": row["categorie_charge"],
                    "rubrique_facture": row[self.col("rubrique_facture")],
                    "type_charge": row[self.col("type_charge")],
                    "libelle": row[self.col("libelle")],
                    "montant_ht": montant_val,
                    "montant_ttc": float(row.get(self.col("montant_ttc"), 0) or 0),
                }
            )

        total_ht = float(montants.sum())
        summary = {
            "total_lignes": len(rows),
            "total_montant_ht": round(total_ht, 2),
            "par_type_ligne": {
                key: {"count": value["count"], "montant_ht": round(float(value["montant_ht"]), 2)}
                for key, value in per_type.items()
            },
        }
        return rows, summary

    # ------------------------------------------------------------------ #
    # Internal helpers
    def _filtered_dataframe(
        self,
        *,
        date: str | None = None,
        date_debut: str | None = None,
        date_fin: str | None = None,
        numero_compte: str | None = None,
        numero_facture: int | None = None,
        numero_acces: str | None = None,
        type_acces: List[str] | None = None,
        type_ligne: List[str] | None = None,
        type_charge: List[str] | None = None,
    ) -> pd.DataFrame:
        if self.data is None:
            raise ValueError("Aucune donnée chargée. Utilisez load_csv() d'abord.")

        df = self.data.copy()
        df["__date_obj"] = pd.to_datetime(df[self.col("date")], format="%d/%m/%Y", errors="coerce")
        df["categorie_charge"] = df.apply(self.categorize_charge, axis=1)
        df["type_ligne"] = df.apply(
            lambda r: self.classify_access_type(
                r[self.col("type_acces")], r[self.col("numero_compte")], r[self.col("libelle")]
            ),
            axis=1,
        )

        if date:
            target = pd.to_datetime(date, format="%Y-%m-%d", errors="coerce")
            if pd.isna(target):
                raise ValueError("Format de date invalide. Utilisez YYYY-MM-DD.")
            df = df[df["__date_obj"].dt.date == target.date()]

        if date_debut:
            start = pd.to_datetime(date_debut, format="%Y-%m-%d", errors="coerce")
            if pd.isna(start):
                raise ValueError("Format de date_debut invalide. Utilisez YYYY-MM-DD.")
            df = df[df["__date_obj"] >= start]

        if date_fin:
            end = pd.to_datetime(date_fin, format="%Y-%m-%d", errors="coerce")
            if pd.isna(end):
                raise ValueError("Format de date_fin invalide. Utilisez YYYY-MM-DD.")
            df = df[df["__date_obj"] <= end]

        if numero_compte:
            df = df[df[self.col("numero_compte")] == numero_compte]
        if numero_facture:
            df = df[df[self.col("numero_facture")] == numero_facture]
        if numero_acces:
            df = df[df[self.col("numero_acces")] == numero_acces]

        def _match_list(values: List[str] | None, series: pd.Series) -> pd.Series:
            if not values:
                return pd.Series([True] * len(series), index=series.index)
            lowered = {v.lower() for v in values}
            return series.fillna("").str.lower().isin(lowered)

        df = df[_match_list(type_acces, df[self.col("type_acces")])]
        df = df[_match_list(type_charge, df[self.col("type_charge")])]
        df = df[_match_list(type_ligne, df["type_ligne"])]

        return df

    def col(self, key: str) -> str:
        return self.columns[key]

    def _resolve_columns(self) -> None:
        """Map known columns to normalized keys for safer access."""

        if self.data is None:
            return

        mapping: Dict[str, str] = {}
        for key, aliases in self.COLUMN_ALIASES.items():
            for alias in aliases:
                if alias in self.data.columns:
                    mapping[key] = alias
                    break
            else:
                raise KeyError(f"Colonne attendue manquante: {key}")
        self.columns = mapping

    def _mark_accounts_with_fiber(self) -> None:
        """Detect accounts that have fibre in any libellé to disambiguate internet vs bas débit."""

        if self.data is None:
            self.accounts_with_fiber = set()
            return

        comp_col = self.col("numero_compte")
        lib_col = self.col("libelle")
        access_col = self.col("type_acces")
        fiber_keywords = ["fibre", "fiber", "ftth", "ftto", "fttx", "optique"]
        bas_keywords = ["bas d", "bas debit", "bas-debit", "basdebit", "rtc", "numeris", "t2", "t0", "adsl", "sdsl", "vdsl", "cuivre", "dsl"]

        accounts: set[str] = set()
        for account, grp in self.data.groupby(comp_col):
            fiber_hits = 0
            bas_hits = 0
            for _, row in grp.iterrows():
                norm_lib = self._normalize_text(str(row.get(lib_col, "")))
                norm_acc = self._normalize_text(str(row.get(access_col, "")))
                if any(k in norm_lib or k in norm_acc for k in fiber_keywords):
                    fiber_hits += 1
                if any(k in norm_lib or k in norm_acc for k in bas_keywords):
                    bas_hits += 1
            total = len(grp)
            if bas_hits == 0 and (
                (fiber_hits >= 2 and (fiber_hits / max(total, 1)) >= 0.1)
                or (total <= 5 and fiber_hits >= 1)
            ):
                accounts.add(str(account))

        self.accounts_with_fiber = accounts

    def _normalize_numeric_columns(self) -> None:
        """Convert columns with comma decimals into floats."""

        assert self.data is not None
        for col_key in ("montant_ht", "montant_tva", "montant_ttc"):
            col = self.columns.get(col_key)
            if col and col in self.data.columns:
                self.data[col] = self.data[col].astype(str).str.replace(",", ".").astype(float)

        numeric_cols = [
            col for col in self.data.columns if "Quantit" in col or "Nombre de lignes" in col
        ]
        for col in numeric_cols:
            self.data[col] = self.data[col].astype(str).str.replace(",", ".").astype(float)

    @staticmethod
    def _normalize_text(text: str) -> str:
        normalized = (
            text.lower()
            .replace("é", "e")
            .replace("è", "e")
            .replace("ê", "e")
            .replace("ë", "e")
            .replace("à", "a")
            .replace("â", "a")
            .replace("ù", "u")
            .replace("û", "u")
            .replace("ô", "o")
            .replace("ï", "i")
            .replace("î", "i")
        )
        return normalized

    def _print_load_summary(self) -> None:
        if self.data is None:
            return
        print(f"✅ Fichier chargé avec succès: {len(self.data)} lignes")
        print(f"  - {self.data[self.col('numero_compte')].nunique()} comptes de facturation")
        print(f"  - {self.data[self.col('numero_facture')].nunique()} factures")
        print(f"  - {self.data[self.col('numero_acces')].nunique()} numéros d'accès")

    @staticmethod
    def _month_label(date_obj: pd.Timestamp) -> str:
        mapping = {
            "January": "Janvier",
            "February": "Février",
            "March": "Mars",
            "April": "Avril",
            "May": "Mai",
            "June": "Juin",
            "July": "Juillet",
            "August": "Août",
            "September": "Septembre",
            "October": "Octobre",
            "November": "Novembre",
            "December": "Décembre",
        }
        return mapping.get(date_obj.strftime("%B"), date_obj.strftime("%B"))

    @staticmethod
    def _print_invoice(invoice: InvoiceBreakdown) -> None:
        print(f"\n  Facture n° {invoice.numero_facture} - {invoice.mois}")
        print(f"    Abonnements : {invoice.montants['Abo']:>10.2f} € HT")
        print(f"    Consommations : {invoice.montants['Conso']:>10.2f} € HT")
        print(f"    Remises : {invoice.montants['Remise']:>10.2f} € HT")
        print("    ----------------------------")
        print(f"    TOTAL : {invoice.total:>10.2f} € HT")
        print(f"    TOTAL TTC : {invoice.total_ttc:>10.2f} € TTC")

    @staticmethod
    def _write_breakdown_rows(ws, row: int, facture: InvoiceBreakdown) -> int:
        """Insert breakdown lines (Abo / Conso / Remise) in the Excel sheet."""

        row += 1
        for label in ["Abo", "Conso", "Remise"]:
            value = facture.montants.get(label, 0.0)
            if value:
                ws.cell(row, 2, label)
                ws.cell(row, 4, label)
                ws.cell(row, 5, round(value, 2))
                row += 1
        return row
