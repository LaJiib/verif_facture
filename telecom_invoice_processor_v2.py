#!/usr/bin/env python3
"""
Outil de vérification des factures télécoms - Version adaptée
Traitement des données de facturation Orange Business avec agrégation par compte
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Tuple
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


class TelecomInvoiceProcessorV2:
    """Processeur pour les factures télécoms - Version adaptée au format Orange"""
    
    def __init__(self):
        self.data = None
        self.aggregated_data = None
        self.accounts_data = {}
    
    def load_csv(self, filepath: str) -> pd.DataFrame:
        """
        Charge le fichier CSV de facturation
        
        Args:
            filepath: Chemin vers le fichier CSV
            
        Returns:
            DataFrame pandas avec les données chargées
        """
        # Le fichier utilise des points-virgules comme séparateur
        self.data = pd.read_csv(filepath, sep=';', encoding='utf-8-sig')
        
        # Convertir les colonnes de montants en float (gérer les virgules comme séparateurs décimaux)
        for col in ['Montant (€ HT)', 'Montant TVA', 'Montant (€ TTC)']:
            if col in self.data.columns:
                # Remplacer les virgules par des points et convertir en float
                self.data[col] = self.data[col].astype(str).str.replace(',', '.').astype(float)
        
        # Convertir les autres colonnes numériques
        numeric_cols = [col for col in self.data.columns if 'Quantité' in col or 'Nombre de lignes' in col]
        for col in numeric_cols:
            if col in self.data.columns:
                self.data[col] = self.data[col].astype(str).str.replace(',', '.').astype(float)
        
        print(f"✓ Fichier chargé avec succès: {len(self.data)} lignes")
        print(f"  - {self.data['Numéro compte'].nunique()} comptes de facturation")
        print(f"  - {self.data['Numéro facture'].nunique()} factures")
        print(f"  - {self.data['Numéro accès'].nunique()} numéros d'accès")
        return self.data
    
    def display_structure(self):
        """Affiche la structure du fichier chargé"""
        if self.data is None:
            print("❌ Aucune donnée chargée")
            return
        
        print("\n=== Structure des données ===")
        print(f"Nombre de lignes: {len(self.data)}")
        print(f"Nombre de colonnes: {len(self.data.columns)}")
        
        print(f"\n📊 Répartition par rubrique:")
        print(self.data['Rubrique facture'].value_counts())
        
        print(f"\n📊 Répartition par type de charge:")
        print(self.data['Type de charge'].value_counts())
        
        print(f"\n💰 Montant total HT: {self.data['Montant (€ HT)'].sum():.2f} €")
        print(f"💰 Montant total TTC: {self.data['Montant (€ TTC)'].sum():.2f} €")
    
    def categorize_charge(self, row) -> str:
        """
        Catégorise une ligne en Abo, Conso ou Remise
        
        Args:
            row: Ligne du DataFrame
            
        Returns:
            Catégorie: 'Abo', 'Conso', ou 'Remise'
        """
        type_charge = row['Type de charge']
        rubrique = row['Rubrique facture']
        
        # Les remises
        if 'remise' in str(rubrique).lower() or 'Remises' in str(type_charge):
            return 'Remise'
        
        # Les abonnements
        if 'Abonnements' in str(type_charge) or 'abonnements et options' in str(rubrique).lower():
            return 'Abo'
        
        # Les consommations
        if 'Consommations' in str(type_charge) or 'consommations' in str(rubrique).lower():
            return 'Conso'
        
        # Par défaut, considérer comme consommation
        return 'Conso'
    
    def aggregate_by_account(self) -> Dict:
        """
        Agrège les données par compte de facturation
        
        Returns:
            Dictionnaire avec les données agrégées par compte
        """
        if self.data is None:
            raise ValueError("Aucune donnée chargée. Utilisez load_csv() d'abord.")
        
        # Ajouter la catégorie à chaque ligne
        self.data['Catégorie'] = self.data.apply(self.categorize_charge, axis=1)
        
        aggregated = {}
        
        # Grouper par compte de facturation
        for account, account_group in self.data.groupby('Numéro compte'):
            # Obtenir les informations de facture
            factures = account_group['Numéro facture'].unique()
            
            account_data = {
                'compte': int(account),
                'factures': {},
                'lignes_telecom': account_group['Numéro accès'].nunique()
            }
            
            # Pour chaque facture de ce compte
            for facture_num in factures:
                facture_data = account_group[account_group['Numéro facture'] == facture_num]
                
                # Extraire la date (mois)
                date_str = facture_data['Date'].iloc[0]
                date_obj = pd.to_datetime(date_str, format='%d/%m/%Y')
                
                # Dictionnaire de traduction des mois
                mois_fr = {
                    'January': 'Janvier', 'February': 'Février', 'March': 'Mars',
                    'April': 'Avril', 'May': 'Mai', 'June': 'Juin',
                    'July': 'Juillet', 'August': 'Août', 'September': 'Septembre',
                    'October': 'Octobre', 'November': 'Novembre', 'December': 'Décembre'
                }
                mois_en = date_obj.strftime('%B')
                mois = mois_fr.get(mois_en, mois_en)
                
                # Calculer les montants par catégorie
                montants = {
                    'Abo': facture_data[facture_data['Catégorie'] == 'Abo']['Montant (€ HT)'].sum(),
                    'Conso': facture_data[facture_data['Catégorie'] == 'Conso']['Montant (€ HT)'].sum(),
                    'Remise': facture_data[facture_data['Catégorie'] == 'Remise']['Montant (€ HT)'].sum(),
                }
                
                total = facture_data['Montant (€ HT)'].sum()
                
                account_data['factures'][facture_num] = {
                    'mois': mois,
                    'date': date_obj,
                    'numero_facture': facture_num,
                    'montants': montants,
                    'total': total,
                    'total_ttc': facture_data['Montant (€ TTC)'].sum(),
                    'nb_lignes_detail': len(facture_data),
                    'statut': 'Validée'  # Par défaut
                }
            
            aggregated[account] = account_data
        
        self.aggregated_data = aggregated
        print(f"\n✓ Agrégation terminée: {len(aggregated)} comptes traités")
        return aggregated
    
    def generate_report(self):
        """Génère un rapport textuel des données agrégées"""
        if self.aggregated_data is None:
            print("❌ Aucune donnée agrégée disponible")
            return
        
        print("\n" + "="*80)
        print("RAPPORT DE FACTURATION TÉLÉCOM PAR COMPTE")
        print("="*80)
        
        total_general = 0
        total_comptes = len(self.aggregated_data)
        
        for account_num, account_data in sorted(self.aggregated_data.items()):
            print(f"\n📋 Compte: {account_num}")
            print(f"   Nombre de lignes télécom: {account_data['lignes_telecom']}")
            print(f"   Nombre de factures: {len(account_data['factures'])}")
            
            for facture_num, facture_data in account_data['factures'].items():
                print(f"\n   📄 Facture N°: {facture_num} - {facture_data['mois']}")
                print(f"      Abonnements:    {facture_data['montants']['Abo']:>10.2f} € HT")
                print(f"      Consommations:  {facture_data['montants']['Conso']:>10.2f} € HT")
                print(f"      Remises:        {facture_data['montants']['Remise']:>10.2f} € HT")
                print(f"      ─────────────────────────────")
                print(f"      TOTAL:          {facture_data['total']:>10.2f} € HT")
                print(f"      TOTAL TTC:      {facture_data['total_ttc']:>10.2f} € TTC")
                
                total_general += facture_data['total']
        
        print("\n" + "="*80)
        print(f"TOTAUX GÉNÉRAUX")
        print(f"  Nombre de comptes: {total_comptes}")
        print(f"  Montant total HT: {total_general:.2f} €")
        print("="*80)
    
    def export_to_excel_format(self, output_path: str, month: str = "Novembre"):
        """
        Exporte les données au format Excel attendu (comme l'exemple fourni)
        
        Args:
            output_path: Chemin du fichier Excel de sortie
            month: Nom du mois à exporter
        """
        if self.aggregated_data is None:
            raise ValueError("Aucune donnée agrégée. Utilisez aggregate_by_account() d'abord.")
        
        # Créer un workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Lot 1 Fixe"
        
        # Styles
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # En-têtes ligne 1
        ws.cell(1, 1, 'Compte de facturation').font = header_font
        ws.cell(1, 1).fill = header_fill
        ws.cell(1, 2, 'Nom du compte').font = header_font
        ws.cell(1, 2).fill = header_fill
        ws.cell(1, 3, month).font = header_font
        ws.cell(1, 3).fill = header_fill
        
        # En-têtes ligne 2
        ws.cell(2, 3, 'n° de facture').font = header_font
        ws.cell(2, 3).fill = header_fill
        ws.cell(2, 4, 'Résumé').font = header_font
        ws.cell(2, 4).fill = header_fill
        ws.cell(2, 5, 'Montant €HT').font = header_font
        ws.cell(2, 5).fill = header_fill
        ws.cell(2, 6, 'Statut').font = header_font
        ws.cell(2, 6).fill = header_fill
        
        # Remplir les données
        row = 3
        for account_num, account_data in sorted(self.aggregated_data.items()):
            # Trouver la facture du mois spécifié
            facture_data = None
            facture_num = None
            
            for fnum, fdata in account_data['factures'].items():
                # Comparer en minuscules pour éviter les problèmes de casse
                # Gérer aussi bien le français que l'anglais
                mois_data = fdata['mois'].lower()
                mois_cherche = month.lower()
                
                # Correspondances français-anglais
                correspondances = {
                    'janvier': 'january', 'février': 'february', 'mars': 'march',
                    'avril': 'april', 'mai': 'may', 'juin': 'june',
                    'juillet': 'july', 'août': 'august', 'septembre': 'september',
                    'octobre': 'october', 'novembre': 'november', 'décembre': 'december'
                }
                
                # Vérifier les correspondances
                match = False
                if mois_data == mois_cherche:
                    match = True
                elif mois_cherche in correspondances and correspondances[mois_cherche] == mois_data:
                    match = True
                elif mois_data in correspondances.values() and mois_cherche in [k for k, v in correspondances.items() if v == mois_data]:
                    match = True
                
                if match:
                    facture_data = fdata
                    facture_num = int(fnum)  # Convertir en int
                    break
            
            if facture_data:
                # Ligne principale avec compte et total
                ws.cell(row, 1, int(account_num))
                ws.cell(row, 2, '')  # Nom du compte (pas dans les données sources)
                ws.cell(row, 3, facture_num)
                ws.cell(row, 4, 'Total')
                ws.cell(row, 5, round(facture_data['total'], 2))
                ws.cell(row, 6, facture_data['statut'])
                
                row += 1
                
                # Ligne Abonnements
                if facture_data['montants']['Abo'] != 0:
                    ws.cell(row, 1, None)
                    ws.cell(row, 2, 'Abo')
                    ws.cell(row, 3, None)
                    ws.cell(row, 4, 'Abo')
                    ws.cell(row, 5, round(facture_data['montants']['Abo'], 2))
                    ws.cell(row, 6, None)
                    row += 1
                
                # Ligne Consommations
                if facture_data['montants']['Conso'] != 0:
                    ws.cell(row, 1, None)
                    ws.cell(row, 2, 'Conso')
                    ws.cell(row, 3, None)
                    ws.cell(row, 4, 'Conso')
                    ws.cell(row, 5, round(facture_data['montants']['Conso'], 2))
                    ws.cell(row, 6, None)
                    row += 1
                
                # Ligne Remises
                if facture_data['montants']['Remise'] != 0:
                    ws.cell(row, 1, None)
                    ws.cell(row, 2, 'Remise')
                    ws.cell(row, 3, None)
                    ws.cell(row, 4, 'Remise')
                    ws.cell(row, 5, round(facture_data['montants']['Remise'], 2))
                    ws.cell(row, 6, None)
                    row += 1
        
        # Ajuster les largeurs de colonnes
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        
        # Sauvegarder
        wb.save(output_path)
        print(f"✓ Fichier Excel créé: {output_path}")
        print(f"  - {row - 3} lignes de données écrites")
    
    def export_summary_csv(self, output_path: str):
        """
        Exporte un résumé au format CSV
        
        Args:
            output_path: Chemin du fichier CSV de sortie
        """
        if self.aggregated_data is None:
            raise ValueError("Aucune donnée agrégée disponible")
        
        rows = []
        for account_num, account_data in sorted(self.aggregated_data.items()):
            for facture_num, facture_data in account_data['factures'].items():
                rows.append({
                    'Compte': account_num,
                    'Facture': facture_num,
                    'Mois': facture_data['mois'],
                    'Abonnements_HT': round(facture_data['montants']['Abo'], 2),
                    'Consommations_HT': round(facture_data['montants']['Conso'], 2),
                    'Remises_HT': round(facture_data['montants']['Remise'], 2),
                    'Total_HT': round(facture_data['total'], 2),
                    'Total_TTC': round(facture_data['total_ttc'], 2),
                    'Nb_lignes_telecom': account_data['lignes_telecom'],
                    'Statut': facture_data['statut']
                })
        
        df = pd.DataFrame(rows)
        df.to_csv(output_path, index=False, encoding='utf-8-sig', sep=';')
        print(f"✓ Fichier CSV résumé créé: {output_path}")


def main():
    """Fonction principale"""
    print("="*80)
    print("OUTIL DE VÉRIFICATION DES FACTURES TÉLÉCOMS - JALON 1 (V2)")
    print("="*80)
    print()
    
    # Exemple d'utilisation
    # processor = TelecomInvoiceProcessorV2()
    # processor.load_csv('DonnéesNov2025.csv')
    # processor.display_structure()
    # processor.aggregate_by_account()
    # processor.generate_report()
    # processor.export_to_excel_format('output/Lot_1_Fixe_Nov_2025.xlsx', 'Novembre')
    # processor.export_summary_csv('output/resume_facturation.csv')
    
    print("ℹ️  Pour utiliser cet outil, importez la classe TelecomInvoiceProcessorV2")


if __name__ == "__main__":
    main()